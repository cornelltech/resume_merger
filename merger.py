import os
import csv
import re

import xlsxwriter
from openpyxl import load_workbook
from xtopdf.PDFWriter import PDFWriter

from PyPDF2 import PdfFileMerger, PdfFileReader

dir_path = os.path.dirname(os.path.realpath(__file__))

def make_schedule_xlsx(schedule, xlsx):
    workbook = xlsxwriter.Workbook(xlsx)
    worksheet = workbook.add_worksheet()

    for i,timeslot in enumerate(schedule):
        worksheet.write('A%d' % (i+1), timeslot[2])
        worksheet.write('B%d' % (i+1), timeslot[1])

    workbook.close()

def turn_xlsx_into_pdf(xlsx, pdf):
    workbook = load_workbook(xlsx, guess_types=True, data_only=True)
    worksheet = workbook.active

    pw = PDFWriter(pdf)
    pw.setFont('Courier', 12)

    ws_range = worksheet['A1:H50']
    for row in ws_range:
        s = ''
        for cell in row:
            if cell.value is None:
                s += ' ' * 11
            else:
                s += str(cell.value).rjust(10) + ' '
        pw.writeLine(s)
    pw.savePage()
    pw.close()

def clean_text(s):
    s = s.lower()
    # ref: https://stackoverflow.com/questions/14361556/remove-all-special-characters-in-java
    s = re.sub(
           r"[^a-zA-Z0-9]",
           '',
           s,
       )
    # we want to trim the year in the back because it will be inconsistent
    # and a source of error because its so arbitrary and breakable.
    s = ''.join(
        [ s[0:-4], re.sub(
           r"\d+",
           '',
           s[-4:],
       ) ]
    )
    return s

def clean_filename(filename, extension='.pdf'):
    if filename:
        return clean_text(filename.replace(extension, '')) + extension

def clean_student_row(row):
    """company,netid,first last name, program/year, time, comment

    We don't care about the last two.
    """
    company = row[0].lower()
    if row[1] != 'break':
        raw_student_info = row[1:4]
        student_info = ''.join(raw_student_info)
        student_info = clean_text(student_info)
    else:
        student_info = 'break'

    return company, student_info, row

def read_student_file(filename, extension='.pdf'):
    companies = {}
    with open(filename, 'r') as csvfile:
        reader = csv.reader(csvfile)
        for i, row in enumerate(reader):
            if i is 0:
                continue
            company, student_info, data = clean_student_row(row)
            if student_info != 'break':
                student_info = student_info + extension
            profile = (student_info, data[4], data[2])
            if company in companies:
                companies[company].append(profile)
            else:
                companies[company] = [profile]
    return companies

def write_and_merge_pdfs(company, student_filenames):
    print('creating pdf for', company)
    print student_filenames
    merger = PdfFileMerger(strict=False)

    for student_filename in student_filenames:
        print student_filename
        try:
            merger.append(PdfFileReader(open(student_filename, 'rb'), strict=False))
        except Exception as e:
            print(e)

    merger.write(os.path.join(dir_path, 'dist', company + '.pdf'))

def normalize_data_files(loc=os.path.join(dir_path, 'data')):
    file_list = os.listdir(loc)
    for filename in file_list:
        cleaned_filename = clean_filename(filename)
        os.rename(
            '/'.join([loc, filename]),
            '/'.join([loc, cleaned_filename]),
        )

def list_of_files_to_merge(schedule, student_tuples):
    files = [schedule]
    for t in student_tuples:
        files.append(os.path.join(dir_path, 'data', t[0]))
    return files

if __name__ == '__main__':
    print('-> Starting Script')

    print('-> Normalizing Filenames in ./data')
    normalize_data_files()
    print('-- Done')

    dist = os.path.join(dir_path, 'dist')
    if not os.path.exists(dist):
        os.makedirs(dist)

    filename = os.path.join(dir_path, 'students.csv')
    student_company_mapping = read_student_file(filename)
    for company in student_company_mapping:
        if company == '':
            pass
        else:
            xlsx = os.path.join(dir_path, 'schedule.xlsx')
            make_schedule_xlsx(student_company_mapping[company], xlsx)
            if not os.path.isdir(os.path.join(dir_path, 'schedules')):
                os.mkdir(os.path.join(dir_path, 'schedules'))
            pdf = os.path.join(dir_path, 'schedules', company + '-schedule.pdf')
            turn_xlsx_into_pdf(xlsx, pdf)
            files_to_merge = list_of_files_to_merge(pdf, student_company_mapping[company])
            write_and_merge_pdfs(company, files_to_merge)
