# XLSXtoPDF.py

# Program to convert the data from an XLSX file to PDF.
# Uses the openpyxl library and xtopdf.

# Author: Vasudev Ram - http://jugad2.blogspot.com
# Copyright 2015 Vasudev Ram.

from openpyxl import load_workbook
from PDFWriter import PDFWriter

workbook = load_workbook('schedule.xlsx', guess_types=True, data_only=True)
worksheet = workbook.active

pw = PDFWriter('schedule.pdf')
pw.setFont('Courier', 12)

ws_range = worksheet.iter_rows('A1:H13')
for row in ws_range:
    s = ''
    for cell in row:
        if cell.value is None:
            s += ' ' * 11
        else:
            s += str(cell.value).rjust(10) + ' '
    pw.writeLine(s)
pw.savePage()
pw.close()
